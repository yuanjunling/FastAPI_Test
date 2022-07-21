#coding=utf-8
import random,string
from datetime import datetime,timedelta
from openpyxl import load_workbook
import pandas as pd
from jose import JWTError,jwt
from typing import Optional
from app.curd.User_curd import get_user_by_name, upload_file
from sqlalchemy.orm import Session
from app.api.config.config import verity_password, SECRET_KEY, ALGORITHM, get_password_hash
from app.model.test01.mysql import SessionLocal
from fastapi import Depends, HTTPException, status, Request, Header
from starlette.responses import FileResponse
from app.api.v1.extensions.logger import logger
import time,pandas

def get_db():
    db=SessionLocal()
    try:
        yield db
    finally:
        db.close()

def jwt_authenticate_user(db,username:str,password:str):
    """验证用户"""
    user = get_user_by_name(db=db,name=username)
    if not user:
        return False
    elif not verity_password(plain_password=password,hashed_password=user.password):
        return False
    else:
        return user

def created_access_token(data:dict,expires_delta:Optional[timedelta]=None):
    """创建访问令牌"""
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)

    to_encode.update({"exp":expire})
    encoded_jwt=jwt.encode(claims=to_encode,key=SECRET_KEY,algorithm=ALGORITHM)

    return encoded_jwt

async def jwt_get_current_user(request: Request,db:Session=Depends(get_db),token: Optional[str] = Header(...)):
    """获取当前用户"""
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not Validate credentials",
        headers={"WWW-Authenticate": "Bearer"}
    )
    try:
        payload = jwt.decode(token=token,key=SECRET_KEY,algorithms=[ALGORITHM])
        username =payload.get("sub")
        if username is None:
            raise credentials_exception
        useris = await request.app.state.redis.get(username)
        if not useris and useris != token:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='用户未登录或者登陆token已经失效')
    except JWTError:
        raise credentials_exception
    user = get_user_by_name(db=db,name=username)
    if user is None:
        raise credentials_exception
    return user

async def download_file(result:dict):
    try:
        file = "./File/file_result/" + str(time.time()) + ".xlsx"
        df = pandas.DataFrame(result)
        df.columns = ["账户", "名称", "手机号码", "性别", "密码"]
        df.to_excel(file, index=False)
        return FileResponse(file, filename="user.xlsx")
    except:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='excel文件下载失败')



# 自定义保存文件函数
async def saveRaw(file,file_name):
    try:
        data = pd.read_excel(file)
        datas = ''.join(random.sample(string.ascii_letters + string.digits, 10))
        file_name_data = datas +file_name
        data.to_excel('/FastAPI_Yuan/app/File/file_test' + '/' +file_name_data , index = None)
        return file_name_data
    except:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='excel本地保存失败')


#excel文件上传到数据库
async def excel_data_uploading_db(file_name_data:str,str_time,db):
    try:
        # 打开工作薄与工作表
        wb = load_workbook('./File/file_test/'+file_name_data)
        sheet = wb.get_sheet_by_name('Sheet1')

        # 计算表格数据的有效行数rows
        num = 1
        while 1:
            cell = sheet.cell(row=num, column=1).value
            if cell:
                num = num + 1
            else:
                # print(num)
                break
        rows = num - 1
        #数据上传到数据库
        upload_file(db=db,rows=rows,sheet=sheet)
        logger.debug('excel文件：{0}上传成功 日志时间：{1}'.format(file_name_data,str_time))
    except:
        logger.error('{0}文件上传失败 日志时间：{1}'.format(file_name_data,str_time))
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='文件上传失败')
    return {"code":"200","message":"文件上传成功"}