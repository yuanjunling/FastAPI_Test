#coding=utf-8
from typing import Optional, List
from fastapi.security import OAuth2PasswordRequestForm
from openpyxl import load_workbook
from pydantic import BaseModel
from app.File.file_data.user_data import user_result
from app.api.config.config import get_password_hash, ACCESS_TOKEN_EXPIRE_MINUTES, oauth2_schema
from app.api.public.public import get_db, jwt_authenticate_user, created_access_token, jwt_get_current_user, \
    download_file, saveRaw
from app.curd.User_curd import get_user_by_name, create_user, user_update, delete_user, get_user_by_names, \
    get_user_by_phone, upload_file
from app.model.test01.database import engine, Base
from app.model.test01.models import User
from app.schemas.User_schemas import ReadUser, CreateUser, Token, UpdateUser
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, Request, status, Body, UploadFile, File
from sqlalchemy.orm import Session
from datetime import datetime, timedelta

from app.api.v1.extensions.logger import logger
import traceback
from app.schemas.annotation_api_user import JiekouCanshuJieshi



application=APIRouter(dependencies=[Depends(jwt_get_current_user)])
Base.metadata.create_all(bind=engine)

#获取单个用户信息
@application.get("/jwt/users/me",response_model=ReadUser,summary='获取用户信息')
async def jwt_read_users_me(username:str,db:Session=Depends(get_db)):
    db_user = get_user_by_name(db=db, name=username)
    # logger.info('这是取单个用户信息接口：username={},当前时间戳为：{tiems}', username, tiems=time.time())
    # my_function(0, 0, 0)
    if db_user is None :
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='您输入的数据有误')
    else:
        return db_user

#获取全部用户信息
@application.get("/jwt/users/mes",response_model=List[ReadUser],summary='获取全部用户信息')
async def jwt_read_users_mes(db:Session=Depends(get_db)):
    db_user = get_user_by_names(db=db)
    return db_user

#修改用户信息
@application.put('/jwt/users/put',summary='修改用户信息')
async def user_updates(user_form:UpdateUser,db:Session=Depends(get_db)):
    db_user = get_user_by_name(db=db, name=user_form.username)
    if not db_user:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='您输入的用户不存在')
    try:
        if user_form.password:
            user_form.password = get_password_hash(user_form.password)
    except:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='密码加密失败')

    try:
        user_update(db=db,name=user_form.username,up=user_form)
        return {"Message":"修改用户信息成功"}
    except:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='修改用户信息失败')

#删除用户信息
@application.delete('/jwt/user/delete',summary='删除用户信息',description='接口1描述')
async def get_user_delete(
        username:str,
        db:Session=Depends(get_db)):
    db_user = get_user_by_name(db=db, name=username)
    if not db_user:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='您输入的用户不存在')
    else:
        try:
            delete_user(db=db,name=username)
            return {"Message":"删除用户信息成功"}
        except:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='删除用户信息失败')

@application.get("/jwt/user/download", summary="下载文件")
async def download_files():
    try:
        file=await download_file(user_result)
        return {
            "rode": "200",
            'message': '操作成功',
            'filpath':file.path
        }
    except:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='文件下载失败')

@application.post("/upload_files")
async def upload_files(db:Session=Depends(get_db),files:UploadFile=File(...)):

    contents = await files.read()
    file_name_data=await saveRaw(contents, files.filename)
    now_time = datetime.now()
    str_time = now_time.strftime("%Y-%m-%d %X")  # 格式化时间字符串

    try:
        # 打开工作薄与工作表
        wb = load_workbook('./File/file_test/'+file_name_data)
        sheet = wb.get_sheet_by_name('Sheet1')

        # 计算表格数据的有效行数rows
        num = 1
        while 1:
            cell = sheet.cell(row=num, column=1).value
            if cell:
                num = num + 1
            else:
                # print(num)
                break
        rows = num - 1
        # print(rows)

        # for循环迭代读取xls文件中的每行数据, 从第二行开始因为需要跳过标题行
        # 注意：openpyxl方式表格列标与行标都是从1开始计算的
        for r in range(2, rows + 1):
            usernameData = sheet.cell(row=r, column=1).value
            userData = sheet.cell(row=r, column=2).value
            phoneData = sheet.cell(row=r, column=3).value
            sexData = sheet.cell(row=r, column=4).value
            passwordData = sheet.cell(row=r, column=5).value
            # values = (numData, nameData, ageData, classesData, scoreData)
            # print(values)
            # 添加数据
            if passwordData:
                password_hash = get_password_hash(str(passwordData))
            upload_file(db=db,usernameData=usernameData,userData=userData,phoneData=phoneData,sexData=sexData,passwordData=password_hash)
        logger.debug('excel文件：{}上传成功 日志时间：{}'.format(file_name_data,str_time))
    except:
        logger.debug('{0}文件上传失败 日志时间：{1}'.format(file_name_data,str_time))
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='文件上传失败')
    return {"code":"200","message":"文件上传成功"}
